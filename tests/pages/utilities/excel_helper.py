import os
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image

class ExcelHelper:
    @staticmethod
    def create_default_template(file_path):
        """Creates a default login_data.xlsx template if it doesn't exist."""
        if os.path.exists(file_path):
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "LoginTestData"
        
        # Headers
        headers = ["TestCaseID", "Description", "Username", "Password", "ExpectedResult", "Status", "ActualMessage", "Screenshot"]
        ws.append(headers)
        
        # Default test cases
        # TC01 is the valid login test case. The user can update 'mitesh' and 'mitesh_password' with real credentials in the Excel file.
        # TC02 is the invalid login test case.
        data = [
            ["TC01", "Valid Login Test", "mitesh", "mitesh", "Success", "", "", ""],
            ["TC02", "Invalid Login Test", "invalid_user@emilo.in", "wrong_password", "Failure", "", "", ""]
        ]
        
        for row in data:
            ws.append(row)
            
        # Set column widths and formatting slightly nicer
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 25
        
        wb.save(file_path)

    @staticmethod
    def read_test_data(file_path):
        """Reads test cases from the Excel file and returns them as a list of dicts."""
        if not os.path.exists(file_path):
            ExcelHelper.create_default_template(file_path)
            
        wb = load_workbook(file_path)
        ws = wb.active
        
        test_cases = []
        # Row 1 is the header
        for row in range(2, ws.max_row + 1):
            tc_id = ws.cell(row=row, column=1).value
            if not tc_id:
                continue
            
            test_cases.append({
                "row_num": row,
                "TestCaseID": tc_id,
                "Description": ws.cell(row=row, column=2).value,
                "Username": ws.cell(row=row, column=3).value or "",
                "Password": ws.cell(row=row, column=4).value or "",
                "ExpectedResult": ws.cell(row=row, column=5).value or "Success"
            })
        return test_cases

    @staticmethod
    def update_test_result(file_path, tc_id, status, actual_message, screenshot_path=None):
        """Updates the status, message, and embeds a screenshot for a given TestCaseID."""
        if not os.path.exists(file_path):
            return
            
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Find the row matching the TestCaseID
        target_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == tc_id:
                target_row = row
                break
                
        if target_row is None:
            wb.close()
            raise ValueError(f"TestCaseID {tc_id} not found in Excel file.")
            
        # Update text cells
        ws.cell(row=target_row, column=6).value = status          # Status column (F)
        ws.cell(row=target_row, column=7).value = actual_message   # ActualMessage column (G)
        
        # Embed screenshot if provided and exists
        if screenshot_path and os.path.exists(screenshot_path):
            try:
                img = Image(screenshot_path)
                # Resize the image so it fits nicely in Excel
                img.width = 160
                img.height = 90
                
                # Set row height to accommodate the image
                ws.row_dimensions[target_row].height = 75
                
                # Clear existing drawing/images in that cell area if possible (openpyxl doesn't make it easy to clear single cell images, but writing is fine)
                # Add image to column H (8)
                ws.add_image(img, f"H{target_row}")
                ws.cell(row=target_row, column=8).value = f"Embedded: {os.path.basename(screenshot_path)}"
            except Exception as e:
                ws.cell(row=target_row, column=8).value = f"Error embedding SS: {str(e)}"
                
        wb.save(file_path)
        wb.close()
